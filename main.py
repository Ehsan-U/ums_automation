import json
import time
from selenium.webdriver import Chrome
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as  EC
from selenium.webdriver.common.keys import Keys
import hashlib
from scrapy.selector import Selector
import csv
import beepy


class UMS():
    def __init__(self,config):
        with open(config,'r') as f:
            data = json.load(f)
        self.username = data.get("username")
        self.password = data.get("password")
        self.shift = data.get("shift")
        self.degree_program = data.get("degree_program")
        self.study_program = data.get("study_program")
        self.course = data.get("course")
        self.section = data.get("section")
        self.grade_item = data.get("grade_item")
        self.marks = data.get("marks_type")
        self.students_file = data.get("filename")
        self.mode = data.get("mode")

    def login(self,browser):
        # login
        browser.get("http://ums.ue.edu.pk/")
        browser.maximize_window()
        browser.find_element(by=By.ID,value="txtUsername").send_keys(self.username)
        browser.find_element(by=By.ID,value="txtPassword").send_keys(self.password)
        browser.find_element(by=By.ID,value="btnLogin").click()
        return browser

    def course_selection(self,browser,students,finished):
        # content iframe
        frmmain = browser.find_element(by=By.XPATH, value="//frameset[@id='frameContent']/frame")
        # menu iframe
        frmmenu = browser.find_element(by=By.XPATH,value="//frameset[@id='Frameset1']/frame[1]")
        # load menu frame
        browser.switch_to.frame(frmmenu)
        browser.find_element(by=By.ID,value="PermissionTreet17").click()
        browser.find_element(by=By.ID,value="PermissionTreet18").click()
        browser.find_element(by=By.ID,value="PermissionTreet20").click()
        # WebDriverWait(driver=browser,timeout=5).until(EC.presence_of_element_located((By.ID,"pcCategories_txtStudentID")))
        time.sleep(2)
        # load default page
        browser.switch_to.default_content()
        # load content frame
        browser.switch_to.frame(frmmain)
        old = ''
        old_ = ''
        for student in students:
            new = student['Roll-no']
            try:
                if new != old:
                    old = new
                    browser.find_element(by=By.ID,value="pcCategories_txtStudentID").send_keys(student['Roll-no'])
                    browser.find_element(by=By.ID,value="pcCategories_btnShow_CD").click()
                    WebDriverWait(driver=browser,timeout=5).until(EC.presence_of_element_located((By.ID,"pcCategories_gvStudyScheme")))
                sel = Selector(text=browser.page_source)  
                for record in sel.xpath("//tr[@class='grdrow']"):
                    code = record.xpath("./td[2]/span/text()").get()
                    add_roll_course = (student['Roll-no']+code).lower().encode('utf-8')
                    temp_id = hashlib.md5(add_roll_course).hexdigest()
                    if temp_id == student['id']:
                        check_element = record.xpath("./td[9]/span/input")
                        # ignore if already checked
                        if check_element.xpath("./@checked").get(): 
                            pass
                        else:
                            # checked it
                            if student['Selected'] == 1:
                                # check
                                browser.find_element(by=By.ID,value=f"{check_element.xpath('./@id').get()}").click()
                                WebDriverWait(driver=browser,timeout=5).until(EC.presence_of_element_located((By.XPATH,"//span[@style='color:Green;' and @class='error']")))
            except Exception as e:
                print(e)
                finished.writerow([student['Name'],student['Roll-no'],'Skipped'])
            else:
                new_ = student['Roll-no']
                if new_ != old_:
                    old_ = new_
                    finished.writerow([student['Name'],student['Roll-no'],'Done'])
        beepy.beep(sound='ready')
        quit = input('Enter q to quit:')
        if quit == "q":
            browser.close()



    def exam_data_entry(self,browser,students,finished):
        # content iframe
        frmmain = browser.find_element(by=By.XPATH, value="//frameset[@id='frameContent']/frame")
        # menu iframe
        frmmenu = browser.find_element(by=By.XPATH, value="//frameset[@id='Frameset1']/frame[1]")
        # load menu frame
        browser.switch_to.frame(frmmenu)
        browser.find_element(by=By.ID,value="PermissionTreet17").click()
        browser.find_element(by=By.ID,value="PermissionTreet23").click()
        # load default page
        browser.switch_to.default_content()
        # load content frame
        browser.switch_to.frame(frmmain)
        # shift
        if self.shift == 'E':
            browser.find_element(by=By.ID, value="rbtnShift_1").click()
        elif self.shift == 'M':
            browser.find_element(by=By.ID, value="rbtnShift_0").click()
        time.sleep(2)
        # select degree program
        browser.find_element(by=By.ID, value="ddlDegreeProgram").click()
        browser.find_element(by=By.XPATH,value=f"//option[@value='{self.degree_program}']").click()
        # select study program
        element = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, f"//option[@value='{self.study_program}']")))
        browser.find_element(by=By.XPATH,value=f"//option[@value='{self.study_program}']").click()
        # select course
        element = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, f"//option[@value='{self.course}']")))
        browser.find_element(by=By.XPATH,value=f"//option[@value='{self.course}']").click()
        # select section
        element = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, f"//option[@value='{self.section}']")))
        browser.find_element(by=By.XPATH,value=f"//option[@value='{self.section}']").click()
        # select grade item
        element = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, f"//option[@value='{self.grade_item}']")))
        browser.find_element(by=By.XPATH,value=f"//option[@value='{self.grade_item}']").click()
        time.sleep(1)
        # click show button
        browser.find_element(by=By.ID,value="btnShow").click()
        time.sleep(2)

        # locating and fill the marks field corresponding to a student
        source = browser.page_source
        sel = Selector(text=source)
        ums_students = browser.find_elements(by=By.XPATH,value="//tr[@class='grdrow']")
        try:
            for row,ums_student in zip(sel.xpath("//tr[@class='grdrow']"),ums_students):
                for roll,name in zip(row.xpath("./td[2]/span/text()").getall(),row.xpath("./td[3]/span/text()").getall()):
                    add_name_roll = name.lower() +' '+roll.lower()
                    id = hashlib.md5(add_name_roll.encode('utf-8')).hexdigest()
                    for student in students:
                        try:
                            if student['id'] == id:
                                ums_student.find_element(by=By.XPATH,value="./td[5]/input").clear()
                                ums_student.find_element(by=By.XPATH,value="./td[5]/input").send_keys(student[f'{self.marks}'])
                                finished.writerow([student['Name'],student['Roll-no'],'Done'])
                                if students[-1]['id'] == id:
                                    ums_student.find_element(by=By.XPATH,value="./td[5]/input").send_keys(Keys.TAB)
                        except Exception as e:
                            print(e)
                            finished.writerow([student['Name'],student['Roll-no'],'Skipped'])
        except Exception as e:
            print(e)
        beepy.beep(sound='ready')
        quit = input('Enter q to quit:')
        if quit == "q":
            browser.close()

    # fill the marks for given students list
    def automate_ums(self,students):
        browser = Chrome(executable_path="/home/lubuntu/ums_automation/chromedriver")
        file = open('finished.csv','a')
        finished = csv.writer(file)
        finished.writerow(['Name','Roll-no','Status'])
        try:
            # login
            browser = self.login(browser)
            if self.mode == 'exam':
                self.exam_data_entry(browser,students,finished)
            elif self.mode == 'course':
                self.course_selection(browser,students,finished)
        except Exception as e:
            print(e)    
            print("\n[+] Quiting..")        
        finally:
            file.close()

    # read excel and return list of student dicts
    def read_excel(self):
        students = []
        wb = openpyxl.load_workbook(filename=self.students_file,read_only=True)
        sheets = wb.sheetnames
        first_row = True
        for sheet in sheets:
            for row in wb[sheet].iter_rows():
                if first_row:
                    first_row = False
                    continue
                else:
                    student = {}
                    # handling file with 7 columns
                    if len(row) == 7:
                        for cell in row:
                            if cell.column == 3:
                                student['Name'] = cell.value
                            elif cell.column == 2:
                                student['Roll-no'] = cell.value
                            elif cell.column == 4:
                                student['Mids'] = cell.value
                            # elif cell.column == 5:
                            #     student['Practical'] = cell.value
                            elif cell.column == 5:
                                student['Sessional'] = cell.value
                            elif cell.column == 7:
                                student['Total'] = cell.value
                        add_name_roll = (student['Name']+' '+student['Roll-no']).lower().encode('utf-8')
                        student['id'] = hashlib.md5(add_name_roll).hexdigest()
                    # handling file with 8 columns
                    elif len(row) == 8:
                        for cell in row:
                            if cell.column == 3:
                                student['Name'] = cell.value
                            elif cell.column == 2:
                                student['Roll-no'] = cell.value
                            elif cell.column == 4:
                                student['Mids'] = cell.value
                            elif cell.column == 5:
                                student['Practical'] = cell.value
                            elif cell.column == 6:
                                student['Sessional'] = cell.value
                            elif cell.column == 8:
                                student['Total'] = cell.value
                        add_name_roll = (student['Name']+' '+student['Roll-no']).lower().encode('utf-8')
                        student['id'] = hashlib.md5(add_name_roll).hexdigest()
                    elif len(row) == 6:
                        for cell in row:
                            if cell.column == 2:
                                student['Name'] = cell.value
                            elif cell.column == 3:
                                student['Roll-no'] = cell.value
                            elif cell.column == 4:
                                student['Course_Code'] = cell.value
                            elif cell.column == 5:
                                student['Course_Title'] = cell.value
                            elif cell.column == 6:
                                student['Selected'] = cell.value
                        add_roll_course = (student['Roll-no']+student['Course_Code']).lower().encode('utf-8')
                        student['id'] = hashlib.md5(add_roll_course).hexdigest()
                    students.append(student)
        return students


ums = UMS('config.json')
# return students dict containing unique id for each student
students = ums.read_excel()
# take students dict and perform marks filling.
ums.automate_ums(students)
